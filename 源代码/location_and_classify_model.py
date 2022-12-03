import numpy as np
from collections import Counter
from openpyxl import load_workbook
from MyFunction import get_row, myfft, get_split_data
from MyModel import MyModel


class location_and_classify_model:
    def __init__(self) -> None:
        self.location_model = MyModel("location.json", 
                                      "location.h5")
        self.classify_model = MyModel("classify.json", 
                                      "classify.h5")
        self.labels = {"攀爬": 0, "大风": 1, "敲击": 2, 
                       "锯磨": 3, "摇晃": 4, "电钻": 5}

    def read_data(self, file_path, start_row:int =0):
        print("正在读取数据")
        data_file = load_workbook(file_path, read_only=True)
        sht = data_file[data_file.sheetnames[0]]
        event = np.array(get_row(sht, 0, sht.max_column,
                                 begin_row=start_row,
                                 rows_end=start_row+500))
        print("读取完毕")
        return event

    def location_and_classify(self, event):
        to_location = []
        for i in range(np.size(event, axis=1)):
            temp = event[:, i]
            to_location.append(np.array(myfft(temp))
                               .reshape(250, 1))
        to_location = np.array(to_location)
        p = self.location_model.predict(to_location)
        # 找到预测中为1的位置
        loc = np.where(p == 1)[0]
        #print(loc)
        
        # 将列表按是否连续的分块
        temp = []
        for i in range(len(loc)):
            if not temp:
                temp.append([loc[i]])
            elif loc[i - 1] + 1 == loc[i]:
                temp[-1].append(loc[i])
            else:
                temp.append([loc[i]])
        #print(temp)

        # 定位结果
        location_results = []
        for l in temp:
            if len(l) >= 12:
                location_results.append([l[0], l[-1]])
        # print(location_results)

        classify_results = []
        for i in range(len(location_results)):
            location_result = location_results[i]
            total_dataset = event[:, location_result[0] : location_result[1] + 1]
            X = []
            to_classify = get_split_data(total_dataset, step_number=8)
            for block in to_classify:
                X.append(block)
            X = np.array(X)
            p = self.classify_model.predict(X)
            # 分类结果
            classify_results.append(Counter(p).
                                    most_common(1)[0][0])
        result = {"labels": classify_results, 
                  "boxes": location_results}
        return result

    def save_result(self, result: dict, file_path, npy_name="result.npy"):
        result['file_path']=file_path
        out={"result":result,"labels":self.labels}
        np.save(npy_name,out)
