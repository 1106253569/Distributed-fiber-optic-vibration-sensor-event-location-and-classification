from datetime import datetime
from multiprocessing import Pool, cpu_count
import numpy as np
from openpyxl import load_workbook
from MyFunction import get_event_json, get_row
from location_and_classify_model import location_and_classify_model
from data_information import data_information


def read_data(file_path, start_row:int =0):
    data_file = load_workbook(file_path, read_only=True)
    sht = data_file[data_file.sheetnames[0]]
    event = np.array(get_row(sht, 0, sht.max_column,
                                begin_row=start_row,
                                rows_end=start_row+500))
    return event


if __name__ == "__main__":
    json_path = "标注.json"
    pool = Pool(cpu_count())
    all_event_information, event_label = get_event_json(json_path)
    dataset_informations = [
        data_information(information) for information in all_event_information
    ]
    beg_time = datetime.now()
    model = location_and_classify_model()

    print("获取预测数据")
    results = [
        pool.apply_async(read_data, [dataset_inf.data_path])
        for dataset_inf in dataset_informations
    ]
    pool.close()
    pool.join()
    print("读取完毕")
    print("共消耗: " + "{:.2f}".format((datetime.now() - beg_time).total_seconds()) + " 秒")
    for n in range(len(results)):
        ans = {
            "labels": dataset_informations[n].data_label,
            "boxes": [dataset_informations[n].event_local],
        }
        res = model.location_and_classify(results[n].get())
        print("---------")
        print("预期:", end="\t")
        print(ans)
        print("结果:", end="\t")
        print(res)
        
