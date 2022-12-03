import numpy as np
from openpyxl import load_workbook
from MyFunction import get_event_json, get_row
from multiprocessing import Pool,cpu_count
from datetime import datetime
from data_information import data_information

# 在****.xlsx中
# 每一列代表着事件发生的位置
# 每一行表示事件发生的时间
# （默认以连续的32行或16行作为一个机器学习或深度学习的事件样本）
# ------------------------------------------
# 在****.json中标记好事件发生类型和位置
# file_name为读取的事件数据文件
# labels为事件的类型
# boxes为人工标记的事件发生位置


def train_data(dataset_inf: data_information):
    data_file = load_workbook(dataset_inf.data_path,
                              read_only=True)
    sht = data_file[data_file.sheetnames[0]]
    n = sht.max_row
    # 读取
    datalist = get_row(
        sht, dataset_inf.event_local[0], 
        dataset_inf.event_local[1], rows_end=n
    )
    dataset = np.array(datalist)
    return dataset


def no_event(dataset_inf: data_information):
    data_file = load_workbook(dataset_inf.data_path, 
                              read_only=True)
    sht = data_file[data_file.sheetnames[0]]
    # 读取
    datalist = get_row(sht, 0, 
                       dataset_inf.event_local[1] + 500, 
                       rows_end=500)
    dataset = np.array(datalist)
    no_event = np.append(
        dataset[:, 0 : dataset_inf.event_local[0]],
        dataset[:, dataset_inf.event_local[1] + 1 :],
        axis=1,
    )
    return no_event


def get_no_event(dataset_informations):
    # 获取非事件列
    pool = Pool(cpu_count())
    print("获取 no_event 中")
    beg = datetime.now()
    results = [
        pool.apply_async(no_event, [dataset_inf])
        for dataset_inf in dataset_informations
    ]
    pool.close()
    pool.join()
    print("读取完毕")
    print("共消耗: " + "{:.2f}".format((datetime.now() - beg).total_seconds()) + " 秒")
    n_temp = np.zeros(500)
    n_temp = n_temp.reshape(500, 1)
    for i in range(len(results)):
        temp = results[i].get()
        path = "/home/T02053124/数据/no_event/" + str(i) + ".csv"
        np.savetxt(path, temp, fmt="%i", delimiter=",")
    print("no_event 保存完毕")


def get_classify_dataset(data_informations):
    # 获取事件列
    pool = Pool(cpu_count())
    print("获取 train_list 中")
    beg = datetime.now()
    results = [
        pool.apply_async(train_data, [dataset_inf]) 
        for dataset_inf in data_informations
    ]
    pool.close()
    pool.join()
    print("读取完毕")
    print("共消耗: " + "{:.2f}".format((datetime.now() - beg).total_seconds()) + " 秒")
    for i in range(len(data_informations)):
        path_name = (
            "/home/T02053124/数据/train_list/"
            + data_informations[i].data_name
            + "_"
            + str(data_informations[i].data_label[0])
            + ".csv"
        )
        np.savetxt(path_name, results[i].get(), fmt="%i", delimiter=",")
    print("train_list 保存完毕")


if __name__ == "__main__":
    # 从json文件中提取事件信息
    json_path = "标注.json"
    all_event_information, event_label = get_event_json(json_path)
    dataset_informations = [
        data_information(information) 
        for information in all_event_information
    ]
    print("当前CPU核心数为{}".format(cpu_count()))
    print("初始化线程池中")
    # 从相应的xlsx文件中读取对应事件的列
    # 存储为csv文件
    get_no_event(dataset_informations)
    get_classify_dataset(dataset_informations)
